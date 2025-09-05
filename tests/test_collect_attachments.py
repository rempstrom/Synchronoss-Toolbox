import importlib
import hashlib
import sys
from pathlib import Path

import pandas as pd
from PIL import Image
from openpyxl import load_workbook


def load_module():
    project_root = Path(__file__).resolve().parents[1]
    sys.path.append(str(project_root))
    return importlib.import_module("synchronoss_parser.collect_attachments")


def test_collect_attachments_copies_files_and_logs_metadata(tmp_path):
    collect_attachments = load_module()

    messages_dir = tmp_path / "messages"
    attachments_dir = messages_dir / "attachments" / "mms" / "in" / "2024-01-01"
    attachments_dir.mkdir(parents=True)

    img1 = Image.new("RGB", (10, 10), color="red")
    exif = img1.getexif()
    exif[274] = 1  # Orientation tag
    img1_path = attachments_dir / "photo1.jpg"
    img1.save(img1_path, exif=exif)

    img2 = Image.new("RGB", (10, 10), color="blue")
    img2_path = attachments_dir / "photo2.png"
    img2.save(img2_path)

    # Contacts lookup mapping numbers to names
    df = pd.DataFrame(
        [
            {"firstname": "Alice", "lastname": "Smith", "phone_numbers": "111"},
            {"firstname": "Bob", "lastname": "Jones", "phone_numbers": "222"},
        ]
    )
    contacts_xlsx = tmp_path / "contacts.xlsx"
    df.to_excel(contacts_xlsx, index=False)

    csv_content = (
        "Date,Type,Direction,Attachments,Body,Sender,Recipients,\"Message ID\"\n"
        "2024-01-01T00:00:00Z,mms,in,photo1.jpg|photo2.png,Hi,111,222,id1\n"
    )
    (messages_dir / "20240101.csv").write_text(csv_content)

    compiled = tmp_path / "Compiled Attachments"
    records, exif_keys = collect_attachments.collect_attachments(
        messages_dir / "attachments", compiled, contacts_xlsx
    )
    logfile = compiled / "log.xlsx"
    collect_attachments.write_excel(records, exif_keys, logfile)

    dest_files = {f.name for f in compiled.iterdir() if f.is_file() and f.name != "log.xlsx"}
    expected_files = {r["File Name"] for r in records}
    assert dest_files == expected_files
    for name in expected_files:
        assert name.startswith("Alice Smith - 2024-01-01 00-00-00")

    md5_1 = hashlib.md5(img1_path.read_bytes()).hexdigest()
    md5_2 = hashlib.md5(img2_path.read_bytes()).hexdigest()

    wb = load_workbook(logfile)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers[:5] == ["File Name", "Date", "Sender", "Recipient", "MD5"]
    assert "Orientation" in headers

    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, r))
        rows[row_data["File Name"]] = row_data

    assert set(rows.keys()) == expected_files

    row_jpg = next(v for k, v in rows.items() if k.endswith(".jpg"))
    assert row_jpg["Date"] == "2024-01-01T00:00:00Z"
    assert row_jpg["Sender"] == "Alice Smith"
    assert row_jpg["Recipient"] == "Bob Jones"
    assert row_jpg["MD5"] == md5_1
    assert row_jpg["Orientation"] == 1

    row_png = next(v for k, v in rows.items() if k.endswith(".png"))
    assert row_png["MD5"] == md5_2
    assert row_png["Sender"] == "Alice Smith"
    assert row_png["Recipient"] == "Bob Jones"
    assert row_png.get("Orientation") in ("", None)


def test_duplicate_filenames_same_message(tmp_path):
    collect_attachments = load_module()

    messages_dir = tmp_path / "messages"
    messages_dir.mkdir()

    attachments_dir = messages_dir / "attachments" / "mms" / "in" / "2024-01-01"
    attachments_dir.mkdir(parents=True)
    img1 = Image.new("RGB", (10, 10), color="red")
    img1.save(attachments_dir / "a.jpg")
    img2 = Image.new("RGB", (10, 10), color="blue")
    img2.save(attachments_dir / "b.jpg")

    csv = (
        "Date,Type,Direction,Attachments,Body,Sender,Recipients,\"Message ID\"\n"
        "2024-01-01T00:00:00Z,mms,in,a.jpg|b.jpg,Hi,Alice,Bob,id1\n"
    )
    (messages_dir / "20240101.csv").write_text(csv)

    compiled = tmp_path / "Compiled Attachments"
    records, _ = collect_attachments.collect_attachments(messages_dir / "attachments", compiled)

    names = sorted(r["File Name"] for r in records)
    assert names == [
        "Alice - 2024-01-01 00-00-00.jpg",
        "Alice - 2024-01-01 00-00-00_1.jpg",
    ]

    assert len(list(compiled.glob("*.jpg"))) == 2
